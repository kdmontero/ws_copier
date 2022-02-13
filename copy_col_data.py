'''
Notes: 
1. Place in the first sheet the table indicating the FCU tagging and the
   corresponding model name. The tagging must be in col A, and the FCU
   model in col B.
2. Place the data of each FCU model in sheet 2 onwards. This will be the
   reference sheets. Note the index of the last sheet.
3. Place the sheets to be filled up after the reference sheets. Note that
   the sheet name/title must be same with the tagging found in the 1st sheet.
4. Place the files in the directory "raw_files" and the find the output files
   in "output" directory.
'''

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

def execute(filename: str) -> None:
    DATA_FIRST_ROW = 8 # starting row to be copied
    DATA_LAST_ROW = 53 # last row to be copied
    DATA_COL = 'D' # column to be copied
    QTY_CELL = 'D11'
    TAGGING_LAST_ROW = 535
    FCU_SHEET_LAST_INDEX = 14
    SAMPLE_CELL = 'D18'

    wb = load_workbook(f'raw_files/{filename}.xlsx')

    print(f'Working on {filename}...')

    def print_worksheet(ws: Worksheet, max_row: int, max_col: int) -> None:
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                char = get_column_letter(col)
                print(ws[char + str(row)].value, end=" ")
            print()

    def copy_column(
            source: Worksheet, 
            target: Worksheet, 
            start_row: int, 
            last_row: int,
            column: str) -> None:
        for row in range(start_row, last_row + 1):
            target[column + str(row)] = source[column + str(row)].value

    # create the tagging dict
    tag_ws = wb.worksheets[0]
    tag = {} # {FCU Tag: FCU model}
    qty = {} # {FCU Tag: FCU qty}
    for i in range(1, TAGGING_LAST_ROW + 1):
        tag[tag_ws['A' + str(i)].value.upper()] = tag_ws['C' + str(i)].value
        qty[tag_ws['A' + str(i)].value.upper()] = tag_ws['B' + str(i)].value

    # save the reference sheets
    ref = {} # {FCU Model: Worksheet}
    for i in range(1, FCU_SHEET_LAST_INDEX):
        ref[wb.worksheets[i].title] = wb.worksheets[i]

    # excecute copy
    for i in range(FCU_SHEET_LAST_INDEX, len(wb.worksheets)):
        if wb.worksheets[i].title.upper() not in tag:
            print(f'{wb.worksheets[i].title} not found')
            continue

        fcu_tagging = tag[wb.worksheets[i].title.upper()]
        source = ref[fcu_tagging]
        copy_column(source, 
                wb.worksheets[i], 
                DATA_FIRST_ROW, 
                DATA_LAST_ROW, 
                DATA_COL)
        wb.worksheets[i][QTY_CELL] = qty[wb.worksheets[i].title.upper()]


    # check if all worksheets are filled up
    for i in range(FCU_SHEET_LAST_INDEX, len(wb.worksheets)):
        if wb.worksheets[i][SAMPLE_CELL] == None:
            print(f'Worksheet {i+1} - {wb.worksheets[i].title} not filled up')

    # check if the tagging is tally with the sheets
    tag_qty = TAGGING_LAST_ROW - 1
    sheets_qty = len(wb.worksheets) - FCU_SHEET_LAST_INDEX
    if tag_qty != sheets_qty:
        print('Items do not tally')
        print(f'No. of items in tagging sheet is {tag_qty}')
        print(f'No. of items in sheets is {sheets_qty}')

    wb.save(f'output/{filename}_out.xlsx')
    print('Done', end='\n\n')


if __name__ == '__main__':
    filenames = ['tower', 'podium', 'basement']

    for filename in filenames:
        execute(filename)
