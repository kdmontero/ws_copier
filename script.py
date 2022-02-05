# Notes: 
# 1. Place the equipment schedule (tagging paired with the correct info) at
#    the first sheet. The tagging must be in col A, and the FCU in col B.
# 2. Place the reference FCU in sheet 2 onwards. Note the sheet index below.
# 3. Place the FCU sheets to be filled up after the reference FCU.

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

DATA_MAX_ROW = 4
DATA_MAX_COL = 2
FCU_REF_SHEET_LAST_INDEX = 4
ES_LAST_ROW = 6
SAMPLE_CELL = 'B2'

wb = load_workbook('FCUScriptingSample.xlsx')

def printsheet(worksheet):
    for row in range(1, DATA_MAX_ROW + 1):
        for col in range(1, DATA_MAX_COL + 1):
            char = get_column_letter(col)
            print(worksheet[char + str(row)].value, end=" ")
        print()
        
def printdata():
    print(len(wb.worksheets))
    for i in range(FCU_REF_SHEET_LAST_INDEX, len(wb.worksheets)):
        print(wb.worksheets[i].title, i)
        printsheet(wb.worksheets[i])
        print('-' * 50)

def copysheet(source, target, max_row, max_col):
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            char = get_column_letter(col)
            target[char + str(row)] = source[char + str(row)].value


# create the equipment sched dict
es_ws = wb['Equipment Sched']
es = {}
for i in range(1, ES_LAST_ROW + 1):
    es[es_ws['A' + str(i)].value] = es_ws['B' + str(i)].value


# save the reference FCU
fcu_ref = {}
for i in range(1, FCU_REF_SHEET_LAST_INDEX):
    fcu_ref[wb.worksheets[i].title] = wb.worksheets[i]


# excecute copy
for i in range(FCU_REF_SHEET_LAST_INDEX, len(wb.worksheets)):
    fcu = es[wb.worksheets[i].title]
    source = fcu_ref[fcu]
    copysheet(source, wb.worksheets[i], DATA_MAX_ROW, DATA_MAX_COL)

wb.save('FCUScriptingSample1.xlsx')
print('done')

# check if all worksheets are filled up
for i in range(FCU_REF_SHEET_LAST_INDEX, len(wb.worksheets)):
    if wb.worksheets[i][SAMPLE_CELL] == None:
        print(f'Worksheet {i + 1} - {wb.worksheets[i].title} did not fill up')

# check if the equipment schedule is tally with the sheets
es_qty = ES_LAST_ROW - 1
sheets_qty = len(wb.worksheets) - FCU_REF_SHEET_LAST_INDEX
if es_qty != sheets_qty:
    print('FCUs do not tally')
    print(f'No. of FCUs in equipment schedule is {es_qty}')
    print(f'No. of FCUs in sheets is {sheets_qty}')
